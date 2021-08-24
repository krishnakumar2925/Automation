import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet: object = wb['sheet1']
    n = int(input("enter the total column where the rectified value should be placed:"))
    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row, 3)
        m = int(input("enter the  percentage to be multiplied: "))
        correct_value = cell.value*m
        correct_cell = sheet.cell(row, 4)
        correct_cell.value = correct_value
    value = Reference(sheet,
                      min_row = 2,
                      max_row = sheet.max_row,
                      min_col = n,
                      max_col = sheet.max_col)
    chart = BarChart()
    b = input("enter the cell to place the chart")
    chart.add_data(value)
    sheet.add_chart(chart, b)

    wb.save(filename)


a = input("enter the name of xlsx file to be modified")
process_workbook(a)
